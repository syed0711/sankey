# -*- coding: utf-8 -*-

import openpyxl
import json
import pandas as pd
import sys

# Get the string argument passed to the script
path_to_file = "file.xlsx"

# laod Workbook with given path string
wb = openpyxl.load_workbook(path_to_file)
sheet = wb.active

# Select specific rows and columns
rows = list(range(281, 289))  # Select rows 1, 3, and 5
cols = list(range(3, 56))

# Create a dataframe from the selected rows and columns
cell_data = []
for row in rows:
    row_data = []
    for col in cols:
        cell_value = sheet.cell(row=row, column=col).value
        row_data.append(cell_value)
    cell_data.append(row_data)
    
row_names = []
for row in rows:
    cell_value = sheet.cell(row=row, column=1).value
    row_names.append(cell_value)
    
#print(row_names)                                                               # debugging purposes
df = pd.DataFrame(cell_data, index=row_names)
df.columns = df.iloc[0]                                                         # Use the values in the first row as column names
df = df. tail(-1)                                                               # Drop the first row as it is being used for column names

previous_column = None
for i, column in enumerate(df.columns):
    if column is None:
        df.columns.values[i] = previous_column
    else:
        previous_column = column

# The XSLX file has column names that repeat, and that are too long, I have to
# manually define the names - horrible & hardcoded (cannot define generic rule)
# both the long names and repetition mess with the Sankey links between nodes
df.iloc[0] = ["Female","Male","18-24","25-34","35-49","50-54","65+","18-34",
              "35+","ABC1","C2DE","White","Asian","Black","Other/Mixed",
              "Refused","Inner London","Outer London","Employed","Unemployed",
              "Full-Time","Part-Time","Student","Retired","Unoccupied",
              "Other[1]","Self-Employed","Private Sector","Public Sector",
              "Charity Sector","Other[2]","Never Worked","<20,000",
              "20,000-39,999","40,000-69,999",">70,000","Unaware","No answer",
              "Own w/ Part Own","Own - Outright","Own - Mortgage",
              "Own - Shared","All Rent Types","Rent - Landlord","Rent - HA & LA",
              "Rent - LA","Rent - HA","Living w/ F&F","Other[3]","Limited",
              "Highly Limited","Mildly Limited","No Disabilities"]

# Print the dataframe
# with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
#     print(df)
# print(df)


# nodes structure
main_labels = list(dict.fromkeys(df.columns))
sub_labels  = list(df.iloc[0])
row_labels  = list(df.index[1:6])
all_labels  = main_labels + sub_labels + row_labels
#print(all_labels)

nodes = [{"node": i, "name": name} for i, name in enumerate(all_labels)]
#print(nodes)

links = []
for idx, item in enumerate(df.iloc[0]):
    value = df.iloc[6, idx]
    col_name = (df == item).idxmax(axis=1)[0]                                  # find the column name based on the value of the first row (sub labels)
    for node in nodes:
        if node["name"] == col_name:
            source_idx = node["node"]
        if node["name"] == item:
            target_idx = node["node"]
    links.append({"source": source_idx,"target": target_idx,"value": value})
    
for idx, item in enumerate(df.iloc[0]):
    for node in nodes:
        if node["name"] == item:
            source_idx = node["node"]
            break
    
    for i in range(0, 5):
        value = round(df.iloc[6, idx]*df.iloc[1+i, idx])
        links.append({"source": source_idx,"target": 64+i,"value": value})
    
json_data = {"nodes": nodes, "links": links}
with open( '/datasets/output.json', "w") as f:                                             # Save JSON to file
    json.dump(json_data, f, indent=2)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    